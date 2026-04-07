from __future__ import annotations

from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


XLSX_PATH = Path("D:/humanities_llm_tutor_project_2026/judge/hand_grade_judge.xlsx")
TARGET_SHEET = "compiled grading"
GRADER_ORDER = ["faizan", "romain", "nishita", "gpt", "claude"]


def main() -> int:
    wb = load_workbook(XLSX_PATH)
    ws = wb[TARGET_SHEET]

    headers = [str(c.value or "").strip() for c in ws[1]]
    header_to_idx = {h: i for i, h in enumerate(headers)}
    persona_idx = header_to_idx["persona type"]
    transcript_idx = header_to_idx["transcript number"]
    grader_idx = header_to_idx["grader name"]

    # Keep transcript groups in first-seen order.
    groups: "OrderedDict[tuple[str, int], list[list[object]]]" = OrderedDict()
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_vals = list(row[: len(headers)])
        persona = str(row_vals[persona_idx] or "").strip().lower()
        try:
            transcript_num = int(row_vals[transcript_idx])
        except (TypeError, ValueError):
            continue
        key = (persona, transcript_num)
        groups.setdefault(key, []).append(row_vals)

    # Rebuild rows with desired grader order inside each transcript group.
    rebuilt_rows: list[list[object]] = []
    order_rank = {g: i for i, g in enumerate(GRADER_ORDER)}
    for rows in groups.values():
        rows_sorted = sorted(
            rows,
            key=lambda r: (
                order_rank.get(str(r[grader_idx] or "").strip().lower(), 999),
                str(r[grader_idx] or "").strip().lower(),
            ),
        )
        rebuilt_rows.extend(rows_sorted)

    # Clear old data rows.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row_vals in rebuilt_rows:
        ws.append(row_vals)

    wb.save(XLSX_PATH)
    print(f"Reordered rows: {len(rebuilt_rows)}")
    print(f"Transcript groups: {len(groups)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
