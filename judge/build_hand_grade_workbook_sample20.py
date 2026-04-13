"""Build judge/hand_grade_workbook.xlsx with 20 stratified transcripts (Claude-score spread).

Layout matches rebuild_hand_grade_workbook.py: faizan/romain/nishita sheets plus compiled grading
with INDEX/MATCH pull-through and =40-SUM(...) totals (rubric_08, 13 deduction columns incl. 1.3.B).
"""

from __future__ import annotations

import importlib.util
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
TRANSCRIPTS = REPO / "transcripts"
OUT_XLSX = REPO / "judge" / "hand_grade_workbook.xlsx"
EXCEL_CELL_MAX = 32767

# (persona_type, transcript_number) — same 20 as stratified list (seed 42).
SAMPLE_20: list[tuple[str, str]] = [
    ("clueless", "0218"),
    ("chaotic", "0097"),
    ("clueless", "0044"),
    ("chaotic", "0085"),
    ("chaotic", "0079"),
    ("clueless", "0025"),
    ("clueless", "0013"),
    ("chaotic", "0015"),
    ("chaotic", "0007"),
    ("clueless", "0242"),
    ("chaotic", "0021"),
    ("clueless", "0297"),
    ("clueless", "0248"),
    ("clueless", "0123"),
    ("chaotic", "0039"),
    ("chaotic", "0198"),
    ("clueless", "0144"),
    ("chaotic", "0062"),
    ("chaotic", "0228"),
    ("chaotic", "0243"),
]

subsections = [
    "1.1.A",
    "1.1.B",
    "1.1.C",
    "1.2.A",
    "1.2.B",
    "1.3.A",
    "1.3.B",
    "2.1.A",
    "2.2.A",
    "2.2.B",
    "3.1.A",
    "3.2.A",
    "3.2.B",
]
compiled_headers = ["persona type", "transcript number", "grader name", *subsections, "total score"]
grader_headers = ["persona type", "transcript number", "transcript", "grader name", *subsections, "total score"]


def format_transcript_cell(data: dict) -> str:
    """Plain-text transcript for hand-grading (one block per exchange)."""
    exchanges = data.get("exchanges")
    if not isinstance(exchanges, list):
        return ""
    blocks: list[str] = []
    for ex in exchanges:
        if not isinstance(ex, dict):
            continue
        turn = ex.get("turn", "")
        tutor = str(ex.get("tutor") or "").strip()
        student = str(ex.get("student") or "").strip()
        reason = str(
            ex.get("pedagogical_reasoning")
            or ex.get("pedagogical_rationale")
            or ex.get("pedagological_reasoning")
            or ""
        ).strip()
        blocks.append(
            f"turn: {turn}\n"
            f"student: {student}\n"
            f"tutor: {tutor}\n"
            f"pedagological reasoning: {reason}"
        )
    text = "\n\n".join(blocks)
    if len(text) > EXCEL_CELL_MAX:
        text = text[: EXCEL_CELL_MAX - 50].rstrip() + "\n\n...[truncated for Excel cell limit]"
    return text


def main() -> int:
    rows: list[tuple[str, str, str]] = []
    for persona_type, tnum in SAMPLE_20:
        tpath = TRANSCRIPTS / persona_type / f"{persona_type}_raw" / f"transcript_{tnum}.json"
        try:
            data = json.loads(tpath.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as e:
            print(f"Failed to read {tpath}: {e}", file=sys.stderr)
            return 1
        body = format_transcript_cell(data)
        rows.append((persona_type, tnum, body))

    rows.sort(key=lambda r: (r[0], int(r[1])))

    wb = Workbook()
    wb.remove(wb.active)
    ws_compiled = wb.create_sheet("compiled grading")
    ws_faizan = wb.create_sheet("faizan grading")
    ws_romain = wb.create_sheet("romain grading")
    ws_nishita = wb.create_sheet("nishita grading")

    header_fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    body_align = Alignment(wrap_text=True, vertical="top")

    def setup_headers(ws, headers: list[str], helper_col_letter: str) -> None:
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
        helper_col = ord(helper_col_letter) - ord("A") + 1
        ws.cell(row=1, column=helper_col, value="_key")
        ws.column_dimensions[helper_col_letter].hidden = True
        ws.freeze_panes = "A2"

    setup_headers(ws_compiled, compiled_headers, "R")
    setup_headers(ws_faizan, grader_headers, "S")
    setup_headers(ws_romain, grader_headers, "S")
    setup_headers(ws_nishita, grader_headers, "S")

    ws_compiled.column_dimensions["A"].width = 14
    ws_compiled.column_dimensions["B"].width = 18
    ws_compiled.column_dimensions["C"].width = 14
    for col in "DEFGHIJKLMNOP":
        ws_compiled.column_dimensions[col].width = 8
    ws_compiled.column_dimensions["Q"].width = 12

    for ws in (ws_faizan, ws_romain, ws_nishita):
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 100
        ws.column_dimensions["D"].width = 14
        for col in "EFGHIJKLMNOPQ":
            ws.column_dimensions[col].width = 8
        ws.column_dimensions["R"].width = 12

    for ws, grader in ((ws_faizan, "faizan"), (ws_romain, "romain"), (ws_nishita, "nishita")):
        row = 2
        for persona_type, tnum, body in rows:
            ws.cell(row=row, column=1, value=persona_type)
            ws.cell(row=row, column=2, value=tnum)
            ccell = ws.cell(row=row, column=3, value=body)
            ccell.alignment = body_align
            ws.cell(row=row, column=4, value=grader)
            ws.cell(row=row, column=18, value=f"=40-SUM(E{row}:Q{row})")
            ws.cell(row=row, column=19, value=f'=A{row}&"|"&B{row}')
            row += 1

    compiled_graders = ["faizan", "romain", "nishita", "claude"]
    row = 2
    for persona_type, tnum, _body in rows:
        for grader in compiled_graders:
            ws_compiled.cell(row=row, column=1, value=persona_type)
            ws_compiled.cell(row=row, column=2, value=tnum)
            ws_compiled.cell(row=row, column=3, value=grader)
            for col_idx in range(4, 17):
                col = get_column_letter(col_idx)
                ws_compiled[f"{col}{row}"] = (
                    f'=IF($C{row}="faizan",IFERROR(INDEX(\'faizan grading\'!$E:$Q,'
                    f'MATCH($A{row}&"|"&$B{row},\'faizan grading\'!$S:$S,0),COLUMN()-3),""),'
                    f'IF($C{row}="romain",IFERROR(INDEX(\'romain grading\'!$E:$Q,'
                    f'MATCH($A{row}&"|"&$B{row},\'romain grading\'!$S:$S,0),COLUMN()-3),""),'
                    f'IF($C{row}="nishita",IFERROR(INDEX(\'nishita grading\'!$E:$Q,'
                    f'MATCH($A{row}&"|"&$B{row},\'nishita grading\'!$S:$S,0),COLUMN()-3),""),"")))'
                )
            ws_compiled.cell(row=row, column=17, value=f"=40-SUM(D{row}:P{row})")
            ws_compiled.cell(row=row, column=18, value=f'=A{row}&"|"&B{row}')
            row += 1

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUT_XLSX)
        out_path = OUT_XLSX
    except PermissionError:
        alt = OUT_XLSX.with_name("hand_grade_workbook_sample20_ready.xlsx")
        wb.save(alt)
        print(
            f"WARNING: {OUT_XLSX.name} is open or locked; wrote {alt.name} instead. "
            "Close Excel and rename or copy over the main workbook.",
            file=sys.stderr,
        )
        out_path = alt
    print(f"Wrote {out_path} ({len(rows)} transcripts; {len(rows) * len(compiled_graders)} compiled rows)")

    fill_spec = importlib.util.spec_from_file_location(
        "fill_claude_hand_workbook", REPO / "judge" / "fill_claude_hand_workbook.py"
    )
    if fill_spec and fill_spec.loader:
        fill_mod = importlib.util.module_from_spec(fill_spec)
        fill_spec.loader.exec_module(fill_mod)
        fr = fill_mod.fill_workbook(out_path)
        if fr != 0:
            return fr
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
