"""Rebuild judge/hand_grade_workbook.xlsx with stratified sample and embedded transcript text."""

from __future__ import annotations

import json
import random
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
TRANSCRIPTS = REPO / "transcripts"
OUT_XLSX = REPO / "judge" / "hand_grade_workbook.xlsx"
RNG = random.Random(42)
EXCEL_CELL_MAX = 32767

subsections = [
    "1.1.A",
    "1.1.B",
    "1.1.C",
    "1.2.A",
    "1.2.B",
    "1.3.A",
    "1.3.B",
    "2.1.A",
    "2.2.A",
    "2.2.B",
    "3.1.A",
    "3.2.A",
    "3.2.B",
]
compiled_headers = ["persona type", "transcript number", "grader name", *subsections, "total score"]
grader_headers = ["persona type", "transcript number", "transcript", "grader name", *subsections, "total score"]


def format_transcript_cell(data: dict) -> str:
    """Plain-text transcript for hand-grading (one block per exchange)."""
    exchanges = data.get("exchanges")
    if not isinstance(exchanges, list):
        return ""
    blocks: list[str] = []
    for ex in exchanges:
        if not isinstance(ex, dict):
            continue
        turn = ex.get("turn", "")
        tutor = str(ex.get("tutor") or "").strip()
        student = str(ex.get("student") or "").strip()
        reason = str(
            ex.get("pedagogical_reasoning")
            or ex.get("pedagogical_rationale")
            or ex.get("pedagological_reasoning")
            or ""
        ).strip()
        blocks.append(
            f"turn: {turn}\n"
            f"student: {student}\n"
            f"tutor: {tutor}\n"
            f"pedagological reasoning: {reason}"
        )
    text = "\n\n".join(blocks)
    if len(text) > EXCEL_CELL_MAX:
        text = text[: EXCEL_CELL_MAX - 50].rstrip() + "\n\n...[truncated for Excel cell limit]"
    return text


def load_by_persona() -> dict[str, list[str]]:
    by_persona: dict[str, list[str]] = defaultdict(list)
    for family in ("chaotic", "cooperative", "clueless"):
        raw_dir = TRANSCRIPTS / family / f"{family}_raw"
        if not raw_dir.is_dir():
            continue
        for path in raw_dir.glob("transcript_*.json"):
            stem = path.relative_to(TRANSCRIPTS).as_posix()[:-5]
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
            except (json.JSONDecodeError, OSError):
                continue
            persona = str(data.get("student_persona", "")).strip()
            if not persona:
                persona = "unknown"
            by_persona[persona].append(stem)
    return by_persona


def sample_family(by_persona: dict[str, list[str]], family: str, n_target: int) -> list[str]:
    keys = sorted(k for k in by_persona if k.startswith(family + "_"))
    selected: list[str] = []
    seen: set[str] = set()
    while len(selected) < n_target:
        progressed = False
        for pk in keys:
            if len(selected) >= n_target:
                break
            pool = [s for s in by_persona[pk] if s not in seen]
            if not pool:
                continue
            selected.append(RNG.choice(pool))
            seen.add(selected[-1])
            progressed = True
        if not progressed:
            break
    if len(selected) < n_target:
        for pk in keys:
            for s in by_persona[pk]:
                if s not in seen and len(selected) < n_target:
                    selected.append(s)
                    seen.add(s)
    return selected[:n_target]


def stem_to_meta(stem: str) -> tuple[str, str, Path]:
    parts = stem.split("/")
    persona_type = parts[0]
    fname = parts[-1]
    m = re.match(r"transcript_(\d+)$", fname)
    num = m.group(1) if m else fname.removeprefix("transcript_")
    path = REPO / "transcripts" / f"{stem}.json"
    return persona_type, num, path


def main() -> int:
    by_persona = load_by_persona()
    selected = (
        sample_family(by_persona, "chaotic", 10)
        + sample_family(by_persona, "cooperative", 10)
        + sample_family(by_persona, "clueless", 10)
    )
    if len(selected) != 30:
        print(f"Expected 30 stems, got {len(selected)}", file=sys.stderr)
        return 1

    rows: list[tuple[str, str, str]] = []
    for stem in selected:
        persona_type, num, tpath = stem_to_meta(stem)
        try:
            data = json.loads(tpath.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as e:
            print(f"Failed to read {tpath}: {e}", file=sys.stderr)
            return 1
        body = format_transcript_cell(data)
        rows.append((persona_type, num, body))

    rows.sort(key=lambda r: (r[0], int(r[1])))

    wb = Workbook()
    wb.remove(wb.active)
    ws_compiled = wb.create_sheet("compiled grading")
    ws_faizan = wb.create_sheet("faizan grading")
    ws_romain = wb.create_sheet("romain grading")
    ws_nishita = wb.create_sheet("nishita grading")

    header_fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    body_align = Alignment(wrap_text=True, vertical="top")

    def setup_headers(ws, headers: list[str], helper_col_letter: str) -> None:
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
        helper_col = ord(helper_col_letter) - ord("A") + 1
        ws.cell(row=1, column=helper_col, value="_key")
        ws.column_dimensions[helper_col_letter].hidden = True
        ws.freeze_panes = "A2"

    setup_headers(ws_compiled, compiled_headers, "R")
    setup_headers(ws_faizan, grader_headers, "S")
    setup_headers(ws_romain, grader_headers, "S")
    setup_headers(ws_nishita, grader_headers, "S")

    ws_compiled.column_dimensions["A"].width = 14
    ws_compiled.column_dimensions["B"].width = 18
    ws_compiled.column_dimensions["C"].width = 14
    for col in "DEFGHIJKLMNOP":
        ws_compiled.column_dimensions[col].width = 8
    ws_compiled.column_dimensions["Q"].width = 12

    for ws in (ws_faizan, ws_romain, ws_nishita):
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 100
        ws.column_dimensions["D"].width = 14
        for col in "EFGHIJKLMNOPQ":
            ws.column_dimensions[col].width = 8
        ws.column_dimensions["R"].width = 12

    for ws, grader in ((ws_faizan, "faizan"), (ws_romain, "romain"), (ws_nishita, "nishita")):
        row = 2
        for persona_type, tnum, body in rows:
            ws.cell(row=row, column=1, value=persona_type)
            ws.cell(row=row, column=2, value=tnum)
            ccell = ws.cell(row=row, column=3, value=body)
            ccell.alignment = body_align
            ws.cell(row=row, column=4, value=grader)
            ws.cell(row=row, column=18, value=f"=40-SUM(E{row}:Q{row})")
            ws.cell(row=row, column=19, value=f'=A{row}&"|"&B{row}')
            row += 1

    compiled_graders = ["faizan", "romain", "nishita", "claude"]
    row = 2
    for persona_type, tnum, _body in rows:
        for grader in compiled_graders:
            ws_compiled.cell(row=row, column=1, value=persona_type)
            ws_compiled.cell(row=row, column=2, value=tnum)
            ws_compiled.cell(row=row, column=3, value=grader)
            for col_idx in range(4, 17):
                col = get_column_letter(col_idx)
                ws_compiled[f"{col}{row}"] = (
                    f'=IF($C{row}="faizan",IFERROR(INDEX(\'faizan grading\'!$E:$Q,'
                    f'MATCH($A{row}&"|"&$B{row},\'faizan grading\'!$S:$S,0),COLUMN()-3),""),'
                    f'IF($C{row}="romain",IFERROR(INDEX(\'romain grading\'!$E:$Q,'
                    f'MATCH($A{row}&"|"&$B{row},\'romain grading\'!$S:$S,0),COLUMN()-3),""),'
                    f'IF($C{row}="nishita",IFERROR(INDEX(\'nishita grading\'!$E:$Q,'
                    f'MATCH($A{row}&"|"&$B{row},\'nishita grading\'!$S:$S,0),COLUMN()-3),""),"")))'
                )
            ws_compiled.cell(row=row, column=17, value=f"=40-SUM(D{row}:P{row})")
            ws_compiled.cell(row=row, column=18, value=f'=A{row}&"|"&B{row}')
            row += 1

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX} ({len(rows)} transcripts)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
