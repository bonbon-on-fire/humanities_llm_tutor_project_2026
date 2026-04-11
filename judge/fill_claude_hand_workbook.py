"""Fill ``compiled grading`` Claude rows in hand_grade_workbook.xlsx from *_claude transcript grades."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
HAND_XLSX = REPO / "judge" / "hand_grade_workbook.xlsx"
MAX_BASE = 40

# Order must match hand-grade workbook deduction columns (rubric_08, no 1.3.C).
SUBSECTION_ORDER = [
    "1.1.A",
    "1.1.B",
    "1.1.C",
    "1.2.A",
    "1.2.B",
    "1.3.A",
    "1.3.B",
    "2.1.A",
    "2.2.A",
    "2.2.B",
    "3.1.A",
    "3.2.A",
    "3.2.B",
]

SUBSECTION_CAPS: dict[str, int] = {
    "1.1.A": 12,
    "1.1.B": 6,
    "1.1.C": 6,
    "1.2.A": 3,
    "1.2.B": 3,
    "1.3.A": 2,
    "1.3.B": 2,
    "2.1.A": 4,
    "2.2.A": 4,
    "2.2.B": 4,
    "3.1.A": 4,
    "3.2.A": 2,
    "3.2.B": 2,
}

# Max total deductions per rubric_08 criterion (parent key).
SECTION_DEDUCTION_CAPS: dict[str, int] = {
    "1.1": 12,
    "1.2": 6,
    "1.3": 2,
    "2.1": 4,
    "2.2": 8,
    "3.1": 4,
    "3.2": 4,
}

SUB_PREFIX = re.compile(r"^(\d+\.\d+\.[A-Z])")


def _coerce_points(raw: object) -> int:
    if isinstance(raw, bool):
        return int(raw)
    if isinstance(raw, int):
        return max(0, raw)
    if isinstance(raw, float):
        return max(0, int(round(raw)))
    if isinstance(raw, str):
        t = raw.strip()
        if not t:
            return 0
        try:
            return max(0, int(round(float(t))))
        except ValueError:
            return 0
    return 0


def _iter_criteria(grade: dict) -> list[dict]:
    out: list[dict] = []
    sections = grade.get("sections")
    if not isinstance(sections, dict):
        return out
    for sec in sections.values():
        if not isinstance(sec, dict):
            continue
        crit = sec.get("criteria")
        if not isinstance(crit, dict):
            continue
        for c in crit.values():
            if isinstance(c, dict):
                out.append(c)
    return out


def subsection_deductions_from_grade(grade: dict) -> dict[str, int]:
    """Map rubric subsection headers (e.g. 1.1.A) to non-negative integer deductions."""
    sums: dict[str, int] = {k: 0 for k in SUBSECTION_ORDER}
    for crit in _iter_criteria(grade):
        for d in crit.get("deductions") or []:
            if not isinstance(d, dict):
                continue
            sid = str(d.get("sub_criterion_id") or "").strip()
            m = SUB_PREFIX.match(sid)
            if not m:
                continue
            key = m.group(1)
            if key not in sums:
                continue
            sums[key] += _coerce_points(d.get("points"))

    for key in SUBSECTION_ORDER:
        cap = SUBSECTION_CAPS.get(key, 0)
        if sums[key] > cap:
            sums[key] = cap

    for parent, cap in SECTION_DEDUCTION_CAPS.items():
        keys = [k for k in SUBSECTION_ORDER if k.startswith(parent + ".")]
        total = sum(sums[k] for k in keys)
        if total <= cap or total == 0:
            continue
        scale = cap / total
        for k in keys:
            sums[k] = int(round(sums[k] * scale))
        drift = cap - sum(sums[k] for k in keys)
        if drift != 0:
            pivot = max(keys, key=lambda x: sums[x])
            sums[pivot] = max(0, sums[pivot] + drift)

    return sums


def claude_transcript_path(persona: str, transcript_num: str) -> Path:
    num = str(transcript_num).strip().zfill(4)
    return REPO / "transcripts" / persona / f"{persona}_claude" / f"transcript_{num}.json"


def fill_workbook(path: Path = HAND_XLSX) -> int:
    if not path.is_file():
        print(f"Missing {path}", file=sys.stderr)
        return 1

    wb = load_workbook(path, data_only=False)
    if "compiled grading" not in wb.sheetnames:
        print("Sheet 'compiled grading' not found.", file=sys.stderr)
        return 1

    ws = wb["compiled grading"]
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if isinstance(val, str) and val.strip():
            headers[val.strip().lower()] = col

    missing = [h for h in SUBSECTION_ORDER if h.lower() not in headers]
    if missing:
        print(f"Workbook missing deduction columns: {missing}", file=sys.stderr)
        return 1

    total_col = headers.get("total score")
    if not total_col:
        print("Missing 'total score' column.", file=sys.stderr)
        return 1

    first_d = min(headers[h.lower()] for h in SUBSECTION_ORDER)
    last_d = max(headers[h.lower()] for h in SUBSECTION_ORDER)
    sum_start = get_column_letter(first_d)
    sum_end = get_column_letter(last_d)

    filled = 0
    mismatches: list[str] = []

    for row in range(2, ws.max_row + 1):
        grader = ws.cell(row=row, column=3).value
        if str(grader or "").strip().lower() != "claude":
            continue

        persona = str(ws.cell(row=row, column=1).value or "").strip()
        t_raw = ws.cell(row=row, column=2).value
        if not persona or t_raw is None:
            continue
        tnum = str(t_raw).strip().zfill(4)

        tpath = claude_transcript_path(persona, tnum)
        if not tpath.is_file():
            print(f"Missing Claude transcript: {tpath}", file=sys.stderr)
            return 1

        data = json.loads(tpath.read_text(encoding="utf-8"))
        grade = data.get("grade")
        if not isinstance(grade, dict):
            print(f"No grade in {tpath}", file=sys.stderr)
            return 1

        subs = subsection_deductions_from_grade(grade)
        for h in SUBSECTION_ORDER:
            col = headers[h.lower()]
            ws.cell(row=row, column=col, value=int(subs[h]))

        ws.cell(row=row, column=total_col, value=f"={MAX_BASE}-SUM({sum_start}{row}:{sum_end}{row})")

        max_b = _coerce_points(grade.get("max_base_score")) or MAX_BASE
        total_b = _coerce_points(grade.get("total_base_score"))
        if total_b == 0 and grade.get("total_score") is not None:
            total_b = _coerce_points(grade.get("total_score"))
        expected_ded = max(0, max_b - total_b)
        actual_ded = sum(int(subs[h]) for h in SUBSECTION_ORDER)
        if actual_ded != expected_ded:
            mismatches.append(
                f"{persona} {tnum}: column deductions sum {actual_ded} vs grade implied {expected_ded} "
                f"(max_base {max_b} - total_base {total_b})"
            )

        filled += 1

    for sheet_name in ("faizan grading", "romain grading", "nishita grading"):
        if sheet_name not in wb.sheetnames:
            continue
        gws = wb[sheet_name]
        for row in range(2, gws.max_row + 1):
            cell = gws.cell(row=row, column=18)
            if isinstance(cell.value, str) and "SUM(E" in cell.value and ":Q" in cell.value:
                cell.value = re.sub(r"=\d+-SUM\(", f"={MAX_BASE}-SUM(", cell.value)

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=total_col)
        if isinstance(cell.value, str) and "SUM(D" in cell.value:
            cell.value = re.sub(r"=\d+-SUM\(", f"={MAX_BASE}-SUM(", cell.value)

    wb.save(path)
    print(f"Updated {filled} Claude rows in {path}")
    if mismatches:
        print("Warnings (deduction columns may need manual check):", file=sys.stderr)
        for m in mismatches:
            print(f"  {m}", file=sys.stderr)
    return 0


def main() -> int:
    return fill_workbook()


if __name__ == "__main__":
    raise SystemExit(main())
